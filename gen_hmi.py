"""Generate Gen4 Spinning HMI BT Protocol Excel workbook."""
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ─── Colour palette ───────────────────────────────────────────────────────────
C_HEADER_BLUE   = "1F4E79"   # dark blue  – sheet headers
C_HEADER_GREEN  = "375623"   # dark green – machine-specific headers
C_HEADER_PURPLE = "44236A"   # purple     – machine-specific sub-headers
C_HEADER_ORANGE = "7F3C00"   # brown-orange – special/unique sections
C_ROW_BLUE_L    = "BDD7EE"   # light blue row
C_ROW_GREEN_L   = "E2EFDA"   # light green row
C_ROW_YELLOW_L  = "FFF2CC"   # light yellow row
C_ROW_ORANGE_L  = "FCE4D6"   # light orange row
C_ROW_PURPLE_L  = "E2D9F3"   # light purple row
C_WHITE         = "FFFFFF"
C_TITLE_GRAY    = "D9D9D9"

# ─── Helpers ──────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(color="000000", size=11):
    return Font(bold=True, color=color, size=size)

def header_font(size=11):
    return Font(bold=True, color="FFFFFF", size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=C_HEADER_BLUE, size=11):
    cell.fill = fill(bg)
    cell.font = header_font(size)
    cell.alignment = center()
    cell.border = thin_border()

def style_sub(cell, bg=C_ROW_BLUE_L):
    cell.fill = fill(bg)
    cell.font = bold_font()
    cell.alignment = left()
    cell.border = thin_border()

def style_data(cell, bg=C_WHITE, bold=False):
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, size=11)
    cell.alignment = left()
    cell.border = thin_border()

def style_code(cell, bg=C_WHITE):
    cell.fill = fill(bg)
    cell.font = Font(name="Courier New", size=10, bold=True)
    cell.alignment = center()
    cell.border = thin_border()

def write_row(ws, row, cols, bg=C_WHITE, bold=False, code_cols=None):
    """Write a data row. code_cols = set of 0-based col indices to style as code."""
    code_cols = code_cols or set()
    for c, val in enumerate(cols):
        cell = ws.cell(row=row, column=c+1, value=val)
        if c in code_cols:
            style_code(cell, bg)
        else:
            style_data(cell, bg, bold)

def section_header(ws, row, title, ncols, bg=C_HEADER_BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=title)
    style_header(cell, bg, size=12)
    return row + 1

def col_headers(ws, row, headers, bg=C_HEADER_BLUE):
    for c, h in enumerate(headers):
        cell = ws.cell(row=row, column=c+1, value=h)
        style_header(cell, bg)
    return row + 1

def set_col_widths(ws, widths):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ─── SHEET 1 : Frame Format ───────────────────────────────────────────────────
def sheet_frame_format(wb):
    ws = wb.create_sheet("Frame Format")
    freeze(ws, "A3")
    set_col_widths(ws, [22, 14, 14, 40, 30])

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    c = ws.cell(r, 1, "Gen4 Spinning — BT RFCOMM SPP Protocol  |  Frame Format & TLV Encoding")
    c.fill = fill("1F4E79"); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    # BT Frame layout
    r = section_header(ws, r, "BT Frame Structure  (all values in hex ASCII, upper-case)", 5, C_HEADER_BLUE)
    r = col_headers(ws, r, ["Field", "Hex Chars", "Bytes", "Value / Description", "Notes"], C_HEADER_BLUE)
    rows = [
        ("SOF",           "2", "1", "7E – Start of Frame",                   "Always 0x7E"),
        ("LENGTH",        "2", "1", "LL – total hex-char count of body",      "body = INFO…EOF inclusive"),
        ("INFO",          "2", "1", "II – opcode / message type",             "See Opcodes sheet"),
        ("SUBSTATE",      "2", "1", "SS – context depends on opcode",         "Machine substate or action code"),
        ("ATTR_COUNT",    "2", "1", "CC – number of TLV fields",              "Parser ignores this; uses position"),
        ("TLV_LIST",      "variable", "variable", "Zero or more TLV fields",  "See TLV Format table below"),
        ("EOF",           "2", "1", "7E – End of Frame",                      "Always 0x7E"),
    ]
    alts = [C_ROW_BLUE_L, C_WHITE]
    for i, row_data in enumerate(rows):
        bg = alts[i % 2]
        write_row(ws, r, row_data, bg, code_cols={0, 1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "Frame Length Calculation", 5, C_HEADER_BLUE)
    note = [
        ("LENGTH field = hex-char count of body",
         "body starts at INFO byte, ends at EOF (7E) inclusive",
         "", "", ""),
        ("Example — Request Settings (no TLVs):",
         "7E | 08 | 03 | 00 | 00 | 7E",
         "", "body = '0300007E' = 8 chars → LENGTH = 08", "Full frame = 7E089900007E → 12 hex chars"),
    ]
    for nd in note:
        write_row(ws, r, nd, C_ROW_YELLOW_L)
        r += 1

    r += 1
    r = section_header(ws, r, "TLV Encoding  (Type-Length-Value)", 5, C_HEADER_BLUE)
    r = col_headers(ws, r, ["TLV Field", "Hex Chars", "Meaning", "Notes", ""], C_HEADER_BLUE)
    tlv_rows = [
        ("TYPE",    "2", "TT – TLV type byte",          "Unique per machine and context", ""),
        ("WIRELEN", "2", "WW – value size in hex chars", "NOT byte count; char count",    "02=1byte, 04=2bytes, 08=4bytes"),
        ("VALUE",   "WW chars", "Actual value encoded", "Big-endian hex",                 ""),
    ]
    for i, row_data in enumerate(tlv_rows):
        write_row(ws, r, row_data, alts[i % 2], code_cols={0, 1})
        r += 1

    r += 1
    r = section_header(ws, r, "TLV Data Types — Wire Encoding", 5, C_HEADER_BLUE)
    r = col_headers(ws, r, ["Data Type", "WireLen (WW)", "Value Hex Chars", "Total TLV Chars", "Example"], C_HEADER_BLUE)
    dtype_rows = [
        ("Compact / UInt8",  "02", "2",  "6 chars",  "Type=0x0B, Value=0x01 → 0B0201"),
        ("UInt16 (Int)",     "04", "4",  "8 chars",  "Type=0x80, Value=80 → 80040050"),
        ("Float (IEEE 754)", "08", "8",  "12 chars", "Type=0x70, Value=80.0f → 700842A00000"),
    ]
    for i, row_data in enumerate(dtype_rows):
        write_row(ws, r, row_data, alts[i % 2], code_cols={1, 2, 3, 4})
        r += 1

    r += 1
    r = section_header(ws, r, "Special Strings (not BT frames — raw ASCII)", 5, C_HEADER_ORANGE)
    r = col_headers(ws, r, ["String", "Hex", "Direction", "Meaning", "Notes"], C_HEADER_ORANGE)
    sp_rows = [
        ("Save Success",    "7E017E", "Board → App", "Settings saved OK",        "6 hex chars, 3 bytes"),
        ("Save Failure",    "7E007E", "Board → App", "Settings save failed",      "6 hex chars, 3 bytes"),
        ("FSC Connect",     "OK+CONN",  "Board → App", "BT device connected",    "Raw ASCII string in RX stream"),
        ("FSC Disconnect",  "OK+LOST",  "Board → App", "BT device disconnected", "Raw ASCII string in RX stream"),
    ]
    for i, row_data in enumerate(sp_rows):
        write_row(ws, r, row_data, alts[i % 2], code_cols={1})
        r += 1

    r += 1
    r = section_header(ws, r, "BT Connection Handshake", 5, C_HEADER_GREEN)
    r = col_headers(ws, r, ["Step", "Direction", "Frame / Action", "Notes", ""], C_HEADER_GREEN)
    hs = [
        ("1", "App detects FSC Connect string", "OK+CONN in RX stream", "Triggers connected state", ""),
        ("2", "App → Board", "Paired From Phone  7E089900007E", "INFO=0x99, SS=0x00, no TLVs", "Wait ~2 s"),
        ("3", "App → Board", "Request Settings  7E080300007E",  "INFO=0x03, SS=0x00, no TLVs", "Board replies with opcode 0x02"),
        ("4", "Board → App", "Settings Frame (opcode 0x02)", "Populated with current EEPROM values", ""),
    ]
    for i, row_data in enumerate(hs):
        write_row(ws, r, row_data, alts[i % 2], code_cols={2})
        r += 1


# ─── SHEET 2 : Opcodes ────────────────────────────────────────────────────────
def sheet_opcodes(wb):
    ws = wb.create_sheet("Opcodes")
    freeze(ws, "A3")
    set_col_widths(ws, [16, 28, 16, 18, 16, 50])

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1, "All Opcodes (INFO byte) — Complete Reference")
    c.fill = fill(C_HEADER_BLUE); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    r = col_headers(ws, r, ["INFO Byte (Hex)", "Name", "Substate Usage", "Direction", "Machines", "Description / Payload"], C_HEADER_BLUE)

    opcodes = [
        ("0x03", "Request Settings",        "0x00 = request",     "App → Board",  "All 4",       "Board replies with opcode 0x02 containing all settings TLVs"),
        ("0x02", "Settings To App",         "0x00",               "Board → App",  "All 4",       "All settings TLVs for the machine. Triggered by 0x03 or after save"),
        ("0x01", "Settings From App",       "0x00=Save, 0x01=Update", "App → Board", "All 4",   "Send settings TLVs. SS=0x00: save to EEPROM. SS=0x01: update RAM only (DF/Flyer/Ring). Carding uses save=true/false"),
        ("0x06", "Machine State",           "0x00–0x05 = substate", "Board → App", "All 4",     "Live status frame — substate, speed, length, sensors, error/pause info + embedded carousel data"),
        ("0x07", "Carousel Request / Info", "motorId in SS field", "Both",          "All 4",     "App→Board: request data for motorId. Board→App: motor telemetry TLVs. Motor ID in substate byte"),
        ("0x04", "Diagnostics",             "0x01=Start, 0x02=CalStop, 0x05=CalStart, 0x06=Stop", "App → Board", "All 4", "Start/stop motor test. Start includes motorId, direction, controlType, speed%, duration TLVs"),
        ("0x05", "Diagnosis Response",      "0x00",               "Board → App",  "All 4",       "Live motor telemetry during diagnosis: RPM, PWM, current, power TLVs"),
        ("0x08", "Gearbox Command",         "0x01=Start, 0x02=Stop, 0x03=SaveLeft, 0x04=SaveRight", "App → Board", "Flyer",  "Gearbox teach-in control. Board replies with opcode 0x09"),
        ("0x09", "Gearbox Response",        "0x00",               "Board → App",  "Flyer",       "TLV 0x01=leftValue, TLV 0x02=rightValue (float or UInt16)"),
        ("0x0A", "Reset Length Counter",    "0x00",               "App → Board",  "All 4",       "Resets length accumulator on board to zero. No TLVs"),
        ("0x0B", "RTF",                     "0x01=On, 0x00=Off",  "App → Board",  "DF, Flyer",   "Running Tension Factor enable / disable toggle. No TLVs"),
        ("0x0C", "CSV Log Control",         "0x01=On, 0x00=Off",  "App → Board",  "All 4",       "Enable/disable board-side logging flag. App also starts/stops file writing locally"),
        ("0x0D", "PID Request",             "0x00",               "App → Board",  "All 4",       "Request PID values for a motor. TLV 0x01 = motorId (UInt8/Compact)"),
        ("0x0E", "PID Response",            "0x00",               "Board → App",  "All 4",       "PID values: TLV 0x01=Kp(float), 0x02=Ki(float), 0x03=FF(float), 0x04=SO(float)"),
        ("0x0F", "PID Update (New)",        "0x00",               "App → Board",  "All 4",       "Send updated PID values. Same TLV codes as 0x0E response"),
        ("0x11", "AL Save / Reset Cuts",    "0x01=AL (DF), 0x00=Flyer", "App → Board", "DF, Flyer", "DF: save AL settings TLVs to EEPROM (opcode 0x11 SS=0x01). Flyer: reset front+back cut counts (SS=0x00, no TLVs)"),
        ("0x12", "AL Response",             "0x00",               "Board → App",  "DF only",     "Board echoes AL settings after GET or SAVE. TLVs: KP(0x20), Sliver6(0x21), Sliver5(0x22), Sliver4(0x23), Target(0x24)"),
        ("0x13", "AL GET",                  "0x01",               "App → Board",  "DF only",     "Request current AL settings from EEPROM. Board replies with 0x12. No TLVs"),
        ("0x15", "AL Calibration Result",   "0x00",               "Board → App",  "DF only",     "End of open-loop calibration: TLV 0x30=avgADC(UInt16), TLV 0x31=sampleCount(UInt16)"),
        ("0x16", "Auto Leveller Toggle",    "0x01=Enable, 0x00=Disable", "App → Board", "DF only", "Enable or disable the auto-leveller hardware. Board saves to EEPROM. No TLVs"),
        ("0x99", "Paired From Phone",       "0x00",               "App → Board",  "All 4",       "First frame sent on BT connect. Tells board app is ready. No TLVs"),
    ]
    alts = [C_ROW_BLUE_L, C_WHITE]
    for i, row in enumerate(opcodes):
        bg = alts[i % 2]
        write_row(ws, r, row, bg, code_cols={0, 2})
        r += 1

    r += 1
    r = section_header(ws, r, "Machine Substates (used in opcode 0x06 SS field)", 6, C_HEADER_GREEN)
    r = col_headers(ws, r, ["Substate (Hex)", "Name", "", "", "", "Description"], C_HEADER_GREEN)
    for i, (code, name, desc) in enumerate([
        ("0x00", "Idle",    "Machine stopped, settings can be changed"),
        ("0x01", "Running", "Machine actively running"),
        ("0x02", "Pause",   "Machine paused — includes pause reason TLVs"),
        ("0x03", "Error",   "Fault state — includes error source + reason TLVs"),
        ("0x04", "Homing",  "Lift homing sequence in progress"),
        ("0x05", "Inching", "Slow manual movement mode"),
    ]):
        write_row(ws, r, [code, name, "", "", "", desc], alts[i % 2], code_cols={0})
        r += 1


# ─── SHEET 3 : Draw Frame ─────────────────────────────────────────────────────
def sheet_draw_frame(wb):
    ws = wb.create_sheet("Draw Frame")
    freeze(ws, "A3")
    set_col_widths(ws, [22, 16, 16, 20, 18, 18, 50])
    NC = 7
    alts = [C_ROW_BLUE_L, C_WHITE]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Draw Frame (DF)  — Complete BT Packet Reference")
    c.fill = fill(C_HEADER_BLUE); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    # ── Settings From App ──
    r = section_header(ws, r, "1. Settings From App  |  opcode 0x01  |  substate 0x01  |  App → Board  |  Save to EEPROM", NC, C_HEADER_GREEN)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Default", "Valid Range", "Notes"], C_HEADER_GREEN)
    df_settings = [
        ("Delivery Speed",       "0x70", "04", "UInt16",      "80",   "50 – 300",     "m/min × 1  (integer)"),
        ("Draft",                "0x71", "08", "Float",       "8.0",  "3.0 – 12.0",   "Dimensionless ratio"),
        ("Length Limit",         "0x72", "04", "UInt16",      "400",  "5 – 1250",     "Metres"),
        ("Ramp Up Time",         "0x73", "04", "UInt16",      "6",    "3 – 15",       "Seconds"),
        ("Ramp Down Time",       "0x74", "04", "UInt16",      "6",    "3 – 15",       "Seconds"),
        ("Creel Tension Factor", "0x75", "08", "Float",       "1.0",  "0.25 – 5.0",   "Dimensionless"),
    ]
    for i, row in enumerate(df_settings):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "2. Settings To App  |  opcode 0x02  |  Board → App  |  Same TLVs as above, sent on GET or after save", NC, C_HEADER_GREEN)
    write_row(ws, r, ["Same TLV codes and types as Settings From App (0x70–0x75). Board sends all fields in one frame.", "", "", "", "", "", ""], C_ROW_BLUE_L)
    r += 1

    r += 1
    # ── Run State ──
    r = section_header(ws, r, "3. Machine State  |  opcode 0x06  |  Board → App  |  Live status — sent periodically while running", NC, C_HEADER_GREEN)
    r = section_header(ws, r, "   3a. Running State (SS = 0x01)", NC, C_HEADER_PURPLE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Unit", "Display", "Notes"], C_HEADER_PURPLE)
    df_run = [
        ("Delivery Speed",      "0x0D", "08", "Float",  "m/min", "deliveryMtrsPerMin", "Current delivery roller speed"),
        ("Current Length",      "0x0E", "08", "Float",  "m",     "currentLength",      "Length produced since last reset"),
        ("AL Sensor State",     "0x0F", "04", "UInt16", "—",     "alSensorActive",     "1=Active, 0=Inactive (RUN_SENSOR_TOGGLE_STATE)"),
        ("Motor ID (what info)","0x09", "02", "UInt8",  "—",     "carouselMotorId",    "Present when carousel data is embedded"),
        ("Motor RPM",           "0x07", "04", "UInt16", "rpm",   "carouselData[rpm]",  "Motor RPM — embedded carousel data"),
        ("Motor Current",       "0x06", "08", "Float",  "A",     "carouselData[current]", "Phase current"),
        ("Motor Temp",          "0x04", "04", "UInt16", "°C",    "carouselData[motorTemp]", "Motor winding temperature"),
        ("MOSFET Temp",         "0x05", "04", "UInt16", "°C",    "carouselData[mosfetTemp]", "Inverter MOSFET temperature"),
        ("Output Metres",       "0x08", "08", "Float",  "m",     "carouselData[outputMtrs]", "Motor output metres"),
        ("Total Power",         "0x0A", "08", "Float",  "W",     "carouselData[totalPower]", "Motor total power"),
    ]
    for i, row in enumerate(df_run):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "   3b. Pause State (SS = 0x02)", NC, C_HEADER_PURPLE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Code Values", "Display", "Notes"], C_HEADER_PURPLE)
    df_pause = [
        ("Pause Reason Code", "0x01", "04", "UInt16", "1=User Paused, 2=Creel Sliver Cut", "pauseReason (string)", ""),
        ("Pause Length",      "0x02", "08", "Float",  "metres",                           "pauseLength",           "Length at time of pause"),
    ]
    for i, row in enumerate(df_pause):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "   3c. Error State (SS = 0x03)", NC, C_HEADER_PURPLE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Code Values", "Display", "Notes"], C_HEADER_PURPLE)
    df_err = [
        ("Error Reason", "0x01", "04", "UInt16",
         "2=OverCurrent,4=OverVoltage,8=UnderVoltage,16=MotorThermistor,32=MOSFETThermistor,64=MotorOverTemp,96=SMPSError,97=AckError,128=MOSFETOverTemp,256=EEPROMWrite,512=EEPROMBad,1024=TrackingError",
         "errorReason", ""),
        ("Error Source", "0x02", "04", "UInt16",
         "1=FrontRoller,2=BackRoller,3=Creel,5=Drafting,11=MotherBoard,12=CANBus",
         "errorSource", ""),
    ]
    for i, row in enumerate(df_err):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    # ── AL Settings ──
    r = section_header(ws, r, "4. AL GET  |  opcode 0x13  |  substate 0x01  |  App → Board  |  No TLVs — board replies with opcode 0x12", NC, C_HEADER_ORANGE)
    write_row(ws, r, ["Frame: 7E0813010000  7E   (SS=0x01, no TLVs, 3-second response timeout)", "", "", "", "", "", ""], C_ROW_ORANGE_L, code_cols={0})
    r += 1

    r += 1
    r = section_header(ws, r, "5. AL SAVE  |  opcode 0x11  |  substate 0x01  |  App → Board  |  Saves AL settings to EEPROM", NC, C_HEADER_ORANGE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Valid Range", "Constraint", "Notes"], C_HEADER_ORANGE)
    al_save = [
        ("KP Gain",            "0x20", "08", "Float",  "0 < KP ≤ 1.0", "",                  "Auto-leveller proportional gain"),
        ("Sliver N+1 Thresh",  "0x21", "04", "UInt16", "—",            "Must be < Sliver N", "ADC counts; Sliver6 in code"),
        ("Sliver N Thresh",    "0x22", "04", "UInt16", "—",            "Must be < Sliver N-1","ADC counts; Sliver5 in code"),
        ("Sliver N-1 Thresh",  "0x23", "04", "UInt16", "—",            "Largest of the three","ADC counts; Sliver4 in code"),
        ("Target g/m",         "0x24", "08", "Float",  "4 < target ≤ 6","",                  "Grams-per-metre target"),
    ]
    for i, row in enumerate(al_save):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "6. AL Response  |  opcode 0x12  |  Board → App  |  Board echoes AL settings (after GET or SAVE)", NC, C_HEADER_ORANGE)
    write_row(ws, r, ["Same TLV codes and types as AL SAVE (0x20–0x24). Board sends all 5 fields.", "", "", "", "", "", ""], C_ROW_ORANGE_L)
    r += 1

    r += 1
    r = section_header(ws, r, "7. Auto Leveller Toggle  |  opcode 0x16  |  App → Board  |  No TLVs", NC, C_HEADER_ORANGE)
    av_rows = [
        ("SS=0x01 → Enable AL hardware",  "7E081601007E", "", "", "", "", ""),
        ("SS=0x00 → Disable AL hardware", "7E081600007E", "", "", "", "", ""),
    ]
    for i, row in enumerate(av_rows):
        write_row(ws, r, row, alts[i % 2], code_cols={1})
        r += 1

    r += 1
    r = section_header(ws, r, "8. AL Calibration  |  opcode 0x04  |  App → Board  |  Open-loop ADC calibration", NC, C_HEADER_ORANGE)
    cal_rows = [
        ("Start calibration", "0x04", "SS=0x05", "7E080405007E", "Firmware runs Back Roller open-loop at duty 450", "", ""),
        ("Stop  calibration", "0x04", "SS=0x02", "7E080402007E", "Firmware stops and sends result frame (opcode 0x15)", "", ""),
    ]
    r = col_headers(ws, r, ["Action", "Opcode", "Substate", "Frame (example)", "Notes", "", ""], C_HEADER_ORANGE)
    for i, row in enumerate(cal_rows):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "9. AL Calibration Result  |  opcode 0x15  |  Board → App", NC, C_HEADER_ORANGE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Unit", "Display", "Notes"], C_HEADER_ORANGE)
    cal_res = [
        ("Average ADC Value", "0x30", "04", "UInt16", "ADC counts", "calibration.avgValue", "Mean of all samples"),
        ("Sample Count",      "0x31", "04", "UInt16", "—",          "calibration.count",    "Number of ADC readings taken"),
    ]
    for i, row in enumerate(cal_res):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    # ── Motor IDs ──
    r = section_header(ws, r, "10. Motor IDs (used in Carousel Request SS field and Diagnosis TLV 0x40)", NC, C_HEADER_BLUE)
    r = col_headers(ws, r, ["Motor Name", "ID (Hex)", "", "", "", "", "Notes"], C_HEADER_BLUE)
    df_motors = [
        ("Production (summary)", "0x0A", "", "", "", "", "Carousel production page ID"),
        ("Front Roller",         "0x01", "", "", "", "", ""),
        ("Back Roller",          "0x02", "", "", "", "", ""),
        ("Creel",                "0x03", "", "", "", "", ""),
        ("Drafting",             "0x05", "", "", "", "", ""),
    ]
    for i, row in enumerate(df_motors):
        write_row(ws, r, row, alts[i % 2], code_cols={1})
        r += 1


# ─── SHEET 4 : Carding ────────────────────────────────────────────────────────
def sheet_carding(wb):
    ws = wb.create_sheet("Carding (Blow Card)")
    freeze(ws, "A3")
    set_col_widths(ws, [26, 16, 16, 20, 18, 18, 50])
    NC = 7
    alts = [C_ROW_GREEN_L, C_WHITE]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Carding (Blow Card) — Complete BT Packet Reference")
    c.fill = fill(C_HEADER_GREEN); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    r = section_header(ws, r, "1. Settings From App  |  opcode 0x01  |  substate 0x00=Save  0x01=Update  |  App → Board", NC, C_HEADER_GREEN)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Default", "Valid Range", "Notes"], C_HEADER_GREEN)
    cd_sets = [
        ("Delivery Speed",       "0x80", "08", "Float",  "8.0",  "2.5 – 50.0", "m/min"),
        ("Card Feed Ratio",      "0x81", "08", "Float",  "5.0",  "0.4 – 20.0", "Dimensionless; used in coiler RPM calc"),
        ("Length Limit",         "0x82", "04", "UInt16", "1000", "100 – 1000", "Metres"),
        ("Cylinder Speed",       "0x83", "04", "UInt16", "750",  "300 – 875",  "RPM"),
        ("Beater Speed",         "0x84", "04", "UInt16", "600",  "300 – 650",  "RPM"),
        ("Picker Cylinder Speed","0x85", "04", "UInt16", "600",  "300 – 700",  "RPM"),
        ("Beater Feed",          "0x86", "04", "UInt16", "10",   "1 – 11",     "Feed speed; ×120 = Beater Feed Motor RPM"),
        ("AF Feed",              "0x87", "04", "UInt16", "7",    "1 – 8",      "Auto-feed speed"),
    ]
    for i, row in enumerate(cd_sets):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "2. Settings To App  |  opcode 0x02  |  Board → App  |  Same TLV codes (0x80–0x87)", NC, C_HEADER_GREEN)
    write_row(ws, r, ["All 8 TLVs sent back on request. Delivery Speed and Card Feed Ratio sent as Float; others as UInt16.", "", "", "", "", "", ""], C_ROW_GREEN_L)
    r += 1

    r += 1
    r = section_header(ws, r, "3. Machine State  |  opcode 0x06  |  Board → App", NC, C_HEADER_GREEN)
    r = section_header(ws, r, "   3a. Running State (SS = 0x01)", NC, "2E7D32")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Unit", "Display", "Notes"], "2E7D32")
    cd_run = [
        ("Delivery Speed",      "0x0D", "08", "Float",  "m/min", "deliveryMtrsPerMin", "Current delivery speed"),
        ("Carding Duct Sensor", "0x0B", "02", "UInt8",  "—",     "ductSensor",         "1=Open, 0=Blocked"),
        ("Coiler Sensor",       "0x0C", "02", "UInt8",  "—",     "coilerSensor",       "1=Open, 0=Blocked"),
        ("Motor ID",            "0x09", "02", "UInt8",  "—",     "carouselMotorId",    "Embedded carousel motor ID"),
        ("Motor RPM",           "0x07", "04", "UInt16", "rpm",   "carouselData[rpm]",  ""),
        ("Motor Current",       "0x06", "08", "Float",  "A",     "carouselData[current]", ""),
        ("Motor Temp",          "0x04", "04", "UInt16", "°C",    "carouselData[motorTemp]", ""),
        ("MOSFET Temp",         "0x05", "04", "UInt16", "°C",    "carouselData[mosfetTemp]", ""),
        ("Output Metres",       "0x08", "08", "Float",  "m",     "carouselData[outputMtrs]", ""),
        ("Total Power",         "0x0A", "08", "Float",  "W",     "carouselData[totalPower]", ""),
    ]
    for i, row in enumerate(cd_run):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "   3b. Pause State (SS = 0x02)", NC, "2E7D32")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Code Values", "Display", "Notes"], "2E7D32")
    cd_pause = [
        ("Pause Reason", "0x01", "04", "UInt16", "1=User Paused, 2=Duct Blocked, 3=Coiler Blocked", "pauseReason", ""),
    ]
    for i, row in enumerate(cd_pause):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "   3c. Error State (SS = 0x03)", NC, "2E7D32")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Code Values", "Display", "Notes"], "2E7D32")
    cd_err = [
        ("Error Information", "0x01", "04", "UInt16",
         "2=OverCurrent,4=OverVoltage,8=UnderVoltage,16=MotorThermistor,32=MOSFETThermistor,64=MotorOverTemp,96=SMPSError,128=MOSFETOverTemp,256=EEPROMWrite,512=EEPROMBad,1024=TrackingError",
         "errorInformation", ""),
        ("Error Source",      "0x02", "04", "UInt16",
         "1=Cylinder,2=Beater,3=Cage,4=CylFeed,5=BtrFeed,6=Coiler,7=PickerCyl,8=AFeed,11=MotherBoard,12=CANBus",
         "errorSource", ""),
        ("Error Code",        "0x03", "04", "UInt16", "Raw error code integer", "errorCode", ""),
    ]
    for i, row in enumerate(cd_err):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "4. Internal Parameters (calculated in app, not sent over BT)", NC, C_HEADER_GREEN)
    r = col_headers(ws, r, ["Parameter Name", "Formula", "Source Fields", "Limit (Motor Max)", "", "", "Notes"], C_HEADER_GREEN)
    cd_int = [
        ("Cylinder Motor RPM",      "cylSpeed × 4",                                 "cylSpeed",                  "≤ 3500 RPM (red if exceeded)", "", "", "Gear ratio = 4"),
        ("Beater Motor RPM",        "btrSpeed × 4",                                 "btrSpeed",                  "≤ 3500 RPM (red if exceeded)", "", "", "Gear ratio = 4"),
        ("Cylinder Feed Motor RPM", "cardFeedRatio × 50",                           "cardFeedRatio",             "≤ 1450 RPM", "", "", "Gear ratio = 50"),
        ("Beater Feed Motor RPM",   "btrFeed × 120",                                "btrFeed",                   "≤ 1450 RPM", "", "", "Gear ratio = 120"),
        ("Cage Motor RPM",          "(deliverySpeed × 1000 ÷ 213.63) × 6.91",       "deliverySpeed",             "≤ 1450 RPM", "", "", "tongueGrooveCircumference=213.63, cageGb=6.91"),
        ("Coiler Motor RPM",        "((deliverySpeed×1000×cardFeedRatio)÷194.779)÷1.656×6.91", "deliverySpeed, cardFeedRatio", "≤ 1450 RPM", "", "", "coilerGrooveCirc=194.779, grooveToGbRatio=1.656, coilerGb=6.91"),
    ]
    for i, row in enumerate(cd_int):
        write_row(ws, r, row, alts[i % 2])
        r += 1

    r += 1
    r = section_header(ws, r, "5. Motor IDs (Carousel Request SS and Diagnosis TLV 0x40)", NC, C_HEADER_BLUE)
    r = col_headers(ws, r, ["Motor Name", "ID (Hex)", "", "", "", "", "Notes"], C_HEADER_BLUE)
    cd_motors = [
        ("Production (summary)", "0x0A", "", "", "", "", "Carousel production summary page"),
        ("Cylinder",             "0x01", "", "", "", "", ""),
        ("Beater",               "0x02", "", "", "", "", ""),
        ("Cage",                 "0x03", "", "", "", "", ""),
        ("Cylinder Feed",        "0x04", "", "", "", "", ""),
        ("Beater Feed",          "0x05", "", "", "", "", ""),
        ("Coiler",               "0x06", "", "", "", "", ""),
        ("Picker Cylinder",      "0x07", "", "", "", "", ""),
        ("AF Feed",              "0x08", "", "", "", "", "Auto-feed motor"),
    ]
    for i, row in enumerate(cd_motors):
        write_row(ws, r, row, alts[i % 2], code_cols={1})
        r += 1


# ─── SHEET 5 : Flyer Frame ────────────────────────────────────────────────────
def sheet_flyer(wb):
    ws = wb.create_sheet("Flyer Frame")
    freeze(ws, "A3")
    set_col_widths(ws, [26, 16, 16, 20, 18, 18, 50])
    NC = 7
    alts = [C_ROW_YELLOW_L, C_WHITE]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Flyer Frame — Complete BT Packet Reference")
    c.fill = fill("7F6000"); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    r = section_header(ws, r, "1. Settings From App  |  opcode 0x01  |  substate 0x01=Update  |  App → Board", NC, "7F6000")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Default", "Valid Range", "Notes"], "7F6000")
    fl_sets = [
        ("Spindle Speed",       "0x50", "04", "UInt16", "650",  "500 – 1100",   "RPM"),
        ("Draft",               "0x51", "08", "Float",  "8.8",  "5.0 – 15.0",   "Dimensionless"),
        ("Twist Per Inch",      "0x52", "08", "Float",  "1.4",  "1.0 – 1.6",    "TPI"),
        ("RTF",                 "0x53", "08", "Float",  "1.0",  "0.5 – 1.5",    "Running Tension Factor"),
        ("Layers",              "0x54", "04", "UInt16", "50",   "1 – 100",       "Number of roving layers"),
        ("Max Height",          "0x55", "04", "UInt16", "270",  "250 – 300",    "mm"),
        ("Roving Width",        "0x56", "08", "Float",  "1.2",  "1.0 – 8.0",    "mm"),
        ("Delta Bobbin Dia",    "0x57", "08", "Float",  "1.1",  "0.3 – 2.5",    "mm per layer increment"),
        ("Bare Bobbin Dia",     "0x58", "04", "UInt16", "48",   "46 – 52",      "mm"),
        ("Ramp Up Time",        "0x59", "04", "UInt16", "12",   "5 – 20",       "Seconds"),
        ("Ramp Down Time",      "0x60", "04", "UInt16", "12",   "5 – 20",       "Seconds"),
        ("Change Layer Time",   "0x61", "04", "UInt16", "800",  "200 – 2500",   "ms — time for layer reversal"),
        ("Cone Angle Factor",   "0x62", "08", "Float",  "1.0",  "0.1 – 3.0",    "Dimensionless"),
    ]
    for i, row in enumerate(fl_sets):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "2. Settings To App  |  opcode 0x02  |  Board → App  |  Same TLV codes (0x50–0x62)", NC, "7F6000")
    write_row(ws, r, ["All 13 TLVs returned on request.", "", "", "", "", "", ""], C_ROW_YELLOW_L)
    r += 1

    r += 1
    r = section_header(ws, r, "3. Machine State  |  opcode 0x06  |  Board → App", NC, "7F6000")
    r = section_header(ws, r, "   3a. Running / Homing State (SS = 0x01 / 0x04)", NC, "A07800")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Unit", "Display", "Notes"], "A07800")
    fl_run = [
        ("Left Lift",           "0x01", "08", "Float",  "mm",    "leftLift",           "Left lift motor position"),
        ("Right Lift",          "0x02", "08", "Float",  "mm",    "rightLift",          "Right lift motor position"),
        ("Layer Count",         "0x03", "04", "UInt16", "—",     "layers",             "Current layer number"),
        ("Front Cut Count",     "0x0F", "04", "UInt16", "count", "frontCutCount",      "Front sensor sliver cut event count"),
        ("Back Cut Count",      "0x10", "04", "UInt16", "count", "backCutCount",       "Back sensor sliver cut event count"),
        ("Motor ID",            "0x09", "02", "UInt8",  "—",     "carouselMotorId",    "Embedded carousel motor ID"),
        ("Motor RPM",           "0x07", "04", "UInt16", "rpm",   "carouselData[rpm]",  ""),
        ("Motor Current",       "0x06", "08", "Float",  "A",     "carouselData[current]", ""),
        ("Motor Temp",          "0x04", "04", "UInt16", "°C",    "carouselData[motorTemp]", ""),
        ("MOSFET Temp",         "0x05", "04", "UInt16", "°C",    "carouselData[mosfetTemp]", ""),
        ("Output Metres",       "0x08", "08", "Float",  "m",     "carouselData[outputMtrs]", "Production page only"),
        ("Total Power",         "0x0A", "08", "Float",  "W",     "carouselData[totalPower]", ""),
    ]
    for i, row in enumerate(fl_run):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "   3b. Pause State (SS = 0x02)", NC, "A07800")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Code Values", "Display", "Notes"], "A07800")
    fl_pause = [
        ("Pause Reason", "0x01", "04", "UInt16", "1=UserPaused,2=SliverCut,3=LiftFault,4=BobbinFull", "pauseReason", ""),
        ("Pause Layer",  "0x02", "04", "UInt16", "Layer at time of pause", "pauseLayer", ""),
    ]
    for i, row in enumerate(fl_pause):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "   3c. Error State (SS = 0x03)", NC, "A07800")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Code Values", "Display", "Notes"], "A07800")
    fl_err = [
        ("Error Information", "0x01", "04", "UInt16",
         "2=OverCurrent,4=OverVoltage,8=UnderVoltage,16=MotorThermistor,32=MOSFETThermistor,64=MotorOverTemp,96=SMPSError,128=MOSFETOverTemp,256=EEPROMWrite,512=EEPROMBad,1024=TrackingError,2048=EncoderSetup,4096=LiftPosTracking,8192=LiftSyncFail,16384=LiftOutOfBounds",
         "errorInformation", ""),
        ("Error Source", "0x02", "04", "UInt16",
         "1=Flyer,2=Bobbin,3=FrontRoller,4=BackRoller,5=Drafting,6=Winding,7=Lift,8=LiftLeft,9=LiftRight,11=MotherBoard,12=CANBus",
         "errorSource", ""),
        ("Error Code", "0x03", "04", "UInt16", "Raw error code", "errorCode", ""),
        ("Error Layer", "0x04", "04", "UInt16", "Layer at time of error", "errorLayer", ""),
    ]
    for i, row in enumerate(fl_err):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "4. Reset Cut Counts  |  opcode 0x11  |  substate 0x00  |  App → Board  |  No TLVs", NC, C_HEADER_ORANGE)
    rc_rows = [
        ("Frame: 7E081100007E", "Sends INFO=0x11, SS=0x00, no TLVs", "Board resets front+back cut counters to 0", "App also resets local display values to 0 immediately", "", "", ""),
    ]
    for i, row in enumerate(rc_rows):
        write_row(ws, r, row, C_ROW_ORANGE_L, code_cols={0})
        r += 1

    r += 1
    r = section_header(ws, r, "5. Gearbox  |  opcode 0x08 (App→Board)  |  opcode 0x09 (Board→App)", NC, "7F6000")
    r = col_headers(ws, r, ["Substate", "Direction", "Meaning", "Notes", "", "", ""], "7F6000")
    gb_rows = [
        ("0x01", "App → Board", "Start gearbox teach-in", "Motor runs and counts"),
        ("0x02", "App → Board", "Stop gearbox teach-in",  "Board sends opcode 0x09 response with left+right values"),
        ("0x03", "App → Board", "Save Left value",        ""),
        ("0x04", "App → Board", "Save Right value",       ""),
    ]
    for i, row in enumerate(gb_rows):
        write_row(ws, r, [row[0], row[1], row[2], row[3], "", "", ""], alts[i % 2], code_cols={0})
        r += 1
    r += 1
    r = section_header(ws, r, "   Gearbox Response  |  opcode 0x09  |  Board → App", NC, "A07800")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "", "", "Notes"], "A07800")
    write_row(ws, r, ["Left Gearbox Value",  "0x01", "04 or 08", "UInt16 or Float", "", "", "Board sends as UInt16 or Float depending on firmware"], alts[0], code_cols={1, 2, 3})
    r += 1
    write_row(ws, r, ["Right Gearbox Value", "0x02", "04 or 08", "UInt16 or Float", "", "", "App checks wireLen to decode correctly"], alts[1], code_cols={1, 2, 3})
    r += 1

    r += 1
    r = section_header(ws, r, "6. Motor IDs (Carousel Request SS and Diagnosis TLV 0x40)", NC, C_HEADER_BLUE)
    r = col_headers(ws, r, ["Motor Name", "ID (Hex)", "", "", "", "", "Notes"], C_HEADER_BLUE)
    fl_motors = [
        ("Production (summary)", "0x0A", "", "", "", "", "Output metres production page"),
        ("Flyer",                "0x01", "", "", "", "", ""),
        ("Bobbin",               "0x02", "", "", "", "", ""),
        ("Front Roller",         "0x03", "", "", "", "", ""),
        ("Back Roller",          "0x04", "", "", "", "", ""),
        ("Drafting",             "0x05", "", "", "", "", ""),
        ("Winding",              "0x06", "", "", "", "", ""),
        ("Lift (combined)",      "0x07", "", "", "", "", "Diagnosis only"),
        ("Lift Left",            "0x08", "", "", "", "", ""),
        ("Lift Right",           "0x09", "", "", "", "", ""),
    ]
    for i, row in enumerate(fl_motors):
        write_row(ws, r, row, alts[i % 2], code_cols={1})
        r += 1


# ─── SHEET 6 : Ring Doubler ───────────────────────────────────────────────────
def sheet_ring(wb):
    ws = wb.create_sheet("Ring Doubler")
    freeze(ws, "A3")
    set_col_widths(ws, [26, 16, 16, 20, 18, 18, 50])
    NC = 7
    alts = [C_ROW_PURPLE_L, C_WHITE]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Ring Doubler — Complete BT Packet Reference")
    c.fill = fill(C_HEADER_PURPLE); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    r = section_header(ws, r, "1. Settings From App  |  opcode 0x01  |  substate 0x01  |  App → Board", NC, C_HEADER_PURPLE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Default", "Valid Range", "Notes"], C_HEADER_PURPLE)
    rg_sets = [
        ("Input Yarn (Ne)",       "0x90", "04", "UInt16", "30",     "10 – 80",    "Yarn count Ne"),
        ("Output Yarn Dia",       "0x91", "08", "Float",  "auto",   "0.05 – 1.5", "mm; auto = –0.10284 + 1.592/√(Ne/2)"),
        ("Spindle Speed",         "0x92", "04", "UInt16", "8000",   "6000 – 12000","RPM"),
        ("Twist Per Inch",        "0x93", "04", "UInt16", "20",     "12 – 30",    "TPI"),
        ("Package Height",        "0x94", "04", "UInt16", "150",    "50 – 250",   "mm"),
        ("Dia Build Factor",      "0x95", "08", "Float",  "0.5",    "0.05 – 2.5", "Dimensionless"),
        ("Winding Closeness",     "0x96", "04", "UInt16", "100",    "50 – 200",   "Percentage (%)"),
        ("Winding Offset Coils",  "0x97", "04", "UInt16", "2",      "1 – 5",      "Count"),
    ]
    for i, row in enumerate(rg_sets):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "2. Settings To App  |  opcode 0x02  |  Board → App  |  Same TLV codes (0x90–0x97)", NC, C_HEADER_PURPLE)
    write_row(ws, r, ["All 8 TLVs returned on request.", "", "", "", "", "", ""], C_ROW_PURPLE_L)
    r += 1

    r += 1
    r = section_header(ws, r, "3. Machine State  |  opcode 0x06  |  Board → App", NC, C_HEADER_PURPLE)
    r = section_header(ws, r, "   3a. Running State (SS = 0x01)", NC, "6A1A9A")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Unit", "Display", "Notes"], "6A1A9A")
    rg_run = [
        ("Weight",              "0x10", "08", "Float",  "g",     "weight",              "Grams per spindle"),
        ("Motor ID",            "0x09", "02", "UInt8",  "—",     "carouselMotorId",     "Embedded carousel motor ID"),
        ("Motor RPM",           "0x07", "04", "UInt16", "rpm",   "carouselData[rpm]",   ""),
        ("Motor Current",       "0x06", "08", "Float",  "A",     "carouselData[current]",""),
        ("Motor Temp",          "0x04", "04", "UInt16", "°C",    "carouselData[motorTemp]",""),
        ("MOSFET Temp",         "0x05", "04", "UInt16", "°C",    "carouselData[mosfetTemp]",""),
        ("Output Metres",       "0x08", "08", "Float",  "m",     "carouselData[outputMtrs]",""),
        ("Total Power",         "0x0A", "08", "Float",  "W",     "carouselData[totalPower]",""),
    ]
    for i, row in enumerate(rg_run):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "   3b. Pause State (SS = 0x02)", NC, "6A1A9A")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Code Values", "Display", "Notes"], "6A1A9A")
    write_row(ws, r, ["Pause Reason", "0x01", "04", "UInt16", "1=UserPaused, 2=CreelSliverCut", "pauseReason", ""], alts[0], code_cols={1, 2, 3})
    r += 1

    r += 1
    r = section_header(ws, r, "   3c. Error State (SS = 0x03)", NC, "6A1A9A")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Code Values", "Display", "Notes"], "6A1A9A")
    rg_err = [
        ("Error Reason", "0x01", "04", "UInt16",
         "2=OverCurrent,4=OverVoltage,8=UnderVoltage,16=MotorThermistor,32=MOSFETThermistor,64=MotorOverTemp,96=SMPSError,97=AckError,98=CanCutError,99=LiftRelPosError,128=MOSFETOverTemp,256=EEPROMWrite,512=EEPROMBad,1024=TrackingError,2048=EncoderSetup,4096=LiftPosTracking,8192=LiftSyncFail,16384=LiftOutOfBounds,32768=EEPROMBadHomingPos",
         "errorReason", ""),
        ("Error Source", "0x02", "04", "UInt16",
         "1=Calender,2=Lift,8=LiftLeft,9=LiftRight,11=MotherBoard,12=CANBus,13=Lifts,14=System",
         "errorSource", ""),
    ]
    for i, row in enumerate(rg_err):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "4. Reset Grams Per Spindle  |  opcode 0x0A  |  substate 0x00  |  App → Board  |  No TLVs", NC, C_HEADER_PURPLE)
    write_row(ws, r, ["Frame: 7E080A00007E", "Resets the weight accumulator on the board back to zero", "", "", "", "", ""], C_ROW_PURPLE_L, code_cols={0})
    r += 1

    r += 1
    r = section_header(ws, r, "5. Output Yarn Diameter Auto-Calculation (App-side formula)", NC, C_HEADER_PURPLE)
    write_row(ws, r, ["Formula: outputYarnDia = –0.10284 + 1.592 / sqrt(inputYarn / 2.0)", "If user leaves field blank, app calculates and sends this value", "", "", "", "", ""], C_ROW_PURPLE_L)
    r += 1

    r += 1
    r = section_header(ws, r, "6. Motor IDs (Carousel Request SS and Diagnosis TLV 0x40)", NC, C_HEADER_BLUE)
    r = col_headers(ws, r, ["Motor Name", "ID (Hex)", "", "", "", "", "Notes"], C_HEADER_BLUE)
    rg_motors = [
        ("Production (summary)", "0x0A", "", "", "", "", ""),
        ("Calender",             "0x01", "", "", "", "", ""),
        ("Lift (combined)",      "0x07", "", "", "", "", "Diagnosis only"),
        ("Lift Left",            "0x08", "", "", "", "", ""),
        ("Lift Right",           "0x09", "", "", "", "", ""),
    ]
    for i, row in enumerate(rg_motors):
        write_row(ws, r, row, alts[i % 2], code_cols={1})
        r += 1


# ─── SHEET 7 : Diagnostics ───────────────────────────────────────────────────
def sheet_diagnostics(wb):
    ws = wb.create_sheet("Diagnostics")
    freeze(ws, "A3")
    set_col_widths(ws, [26, 16, 16, 20, 18, 18, 50])
    NC = 7
    alts = [C_ROW_BLUE_L, C_WHITE]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Diagnosis Protocol — Common to All 4 Machines")
    c.fill = fill(C_HEADER_BLUE); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    r = section_header(ws, r, "Start Diagnosis  |  opcode 0x04  |  substate 0x01  |  App → Board", NC, C_HEADER_BLUE)
    r = section_header(ws, r, "Standard Motor (non-lift)", NC, "274E6A")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Values / Range", "Required", "Notes"], "274E6A")
    diag_start = [
        ("Motor ID",      "0x40", "02", "UInt8",  "Machine-specific motor ID", "Yes", "See Motor IDs on each machine sheet"),
        ("Direction",     "0x44", "02", "UInt8",  "0x01=Forward, 0x02=Reverse", "Yes", ""),
        ("Control Type",  "0x41", "02", "UInt8",  "0x01=OpenLoop, 0x02=ClosedLoop", "Yes", ""),
        ("Speed %",       "0x42", "04", "UInt16", "0 – 100 (percent)",          "Yes", "Not sent for lift motors"),
        ("Duration (s)",  "0x43", "04", "UInt16", "1 – 30 (seconds)",           "Yes", "Not sent for lift motors"),
    ]
    for i, row in enumerate(diag_start):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "Lift Motor (Flyer / Ring lift diagnosis)", NC, "274E6A")
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Values / Range", "Required", "Notes"], "274E6A")
    lift_diag = [
        ("Motor ID",        "0x40", "02", "UInt8",  "0x07=Lift, 0x08=LiftLeft, 0x09=LiftRight", "Yes", ""),
        ("Direction",       "0x44", "02", "UInt8",  "0x01=Up, 0x02=Down",                       "Yes", ""),
        ("Control Type",    "0x41", "02", "UInt8",  "0x01=OpenLoop, 0x02=ClosedLoop",            "Yes", ""),
        ("Bed Distance mm", "0x45", "04", "UInt16", "Distance in mm",                            "Yes", "Replaces Speed% and Duration for lift"),
    ]
    for i, row in enumerate(lift_diag):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "Diagnosis Response  |  opcode 0x05  |  substate 0x00  |  Board → App  (live, while running)", NC, C_HEADER_BLUE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Unit", "Display", "Notes"], C_HEADER_BLUE)
    diag_resp = [
        ("RPM",     "0x01", "04",       "UInt16",       "rpm",  "diagnosisState.rpm",     "Motor speed"),
        ("PWM",     "0x02", "04 or 08", "UInt16 or Float", "%", "diagnosisState.pwm",     "App checks wireLen to decode"),
        ("Current", "0x03", "08",       "Float",        "A",    "diagnosisState.current", "Phase current"),
        ("Power",   "0x04", "08",       "Float",        "W",    "diagnosisState.power",   "Total power"),
    ]
    for i, row in enumerate(diag_resp):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "Stop Diagnosis  |  opcode 0x04  |  substate 0x06  |  App → Board  |  No TLVs", NC, C_HEADER_BLUE)
    stop_rows = [
        ("Frame: 7E080406007E", "Sent 3× consecutively for reliability", "Queue drained (drainTxQueue) before sending stop", "Ensures stale carousel frames do not follow stop command", "", "", ""),
    ]
    for i, row in enumerate(stop_rows):
        write_row(ws, r, row, C_ROW_YELLOW_L, code_cols={0})
        r += 1

    r += 1
    r = section_header(ws, r, "Reliability Notes", NC, C_HEADER_ORANGE)
    notes = [
        "1. drainTxQueue() is called before EVERY diagnosis start and stop — this flushes any pending carousel request frames (0x07) that may have queued up from the logFetchJob background coroutine.",
        "2. Stop command is sent 3× — belt-and-suspenders for intermittent BT loss.",
        "3. logFetchJob pauses carousel requests while isDiagnosing=true to avoid BT queue pollution during diagnosis.",
        "4. On the Tests screen, the carousel LaunchedEffect is disposed (Compose recomposition). But logFetchJob runs in ViewModel scope, so drain + pause are essential.",
    ]
    for i, note in enumerate(notes):
        write_row(ws, r, [note, "", "", "", "", "", ""], alts[i % 2])
        r += 1


# ─── SHEET 8 : Carousel ──────────────────────────────────────────────────────
def sheet_carousel(wb):
    ws = wb.create_sheet("Carousel")
    freeze(ws, "A3")
    set_col_widths(ws, [26, 16, 16, 20, 18, 18, 50])
    NC = 7
    alts = [C_ROW_BLUE_L, C_WHITE]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Carousel Protocol — opcode 0x07 — Common to All 4 Machines")
    c.fill = fill(C_HEADER_BLUE); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    r = section_header(ws, r, "Carousel Request  |  opcode 0x07  |  Motor ID in SUBSTATE field  |  App → Board", NC, C_HEADER_BLUE)
    write_row(ws, r, ["Frame structure: 7E08 07 [motorId] 00 7E   (no TLVs — motorId goes in SS byte, not a TLV)", "", "", "", "", "", ""], C_ROW_YELLOW_L, code_cols={0})
    r += 1
    write_row(ws, r, ["Example (Front Roller, ID=0x01): 7E08070100 7E", "", "", "", "", "", "Trigger: page change or logFetchJob cycling all motors every 400ms"], C_ROW_BLUE_L, code_cols={0})
    r += 1

    r += 1
    r = section_header(ws, r, "Carousel Response  |  opcode 0x07  |  Board → App  — Standard Motor Telemetry TLVs", NC, C_HEADER_BLUE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Unit", "Present in", "Notes"], C_HEADER_BLUE)
    car_tlvs = [
        ("Motor ID (what info)", "0x09", "02", "UInt8",  "—",    "All machines",   "Identifies which motor this frame describes"),
        ("Motor RPM",            "0x07", "04", "UInt16", "rpm",  "All machines",   "Motor rotational speed"),
        ("Motor Current",        "0x06", "08", "Float",  "A",    "All machines",   "Phase current"),
        ("Motor Temp",           "0x04", "04", "UInt16", "°C",   "All machines",   "Motor winding temperature"),
        ("MOSFET Temp",          "0x05", "04", "UInt16", "°C",   "All machines",   "Inverter MOSFET temperature"),
        ("Total Power",          "0x0A", "08", "Float",  "W",    "All machines",   "Motor total power"),
        ("Output Metres",        "0x08", "08", "Float",  "m",    "All machines",   "Production metres (usually Production ID 0x0A page only)"),
        ("Delivery Speed",       "0x0D", "08", "Float",  "m/min","DF, Carding",    "Present in Production carousel of DF and Carding"),
        ("Current Length",       "0x0E", "08", "Float",  "m",    "DF",             "Present in DF Production carousel"),
    ]
    for i, row in enumerate(car_tlvs):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "Production Page Motor IDs per Machine (used in Carousel + logFetchJob cycle)", NC, C_HEADER_GREEN)
    r = col_headers(ws, r, ["Machine", "Production ID", "Motor IDs cycled by logFetchJob", "", "", "", "Notes"], C_HEADER_GREEN)
    prod_rows = [
        ("Draw Frame (DF)",    "0x0A", "0x0A, 0x01, 0x02, 0x03 (Production, FrontRoller, BackRoller, Creel)", "", "", "", ""),
        ("Carding",            "0x0A", "0x0A, 0x01, 0x02, 0x03, 0x04, 0x05, 0x06, 0x07, 0x08 (all 9 motors)", "", "", "", ""),
        ("Flyer Frame",        "0x0A", "0x0A, 0x01, 0x02, 0x03, 0x04, 0x08, 0x09 (7 motors)", "", "", "", ""),
        ("Ring Doubler",       "0x0A", "0x0A, 0x01, 0x08, 0x09 (Production, Calender, LiftLeft, LiftRight)", "", "", "", ""),
    ]
    for i, row in enumerate(prod_rows):
        write_row(ws, r, row, alts[i % 2], code_cols={1})
        r += 1

    r += 1
    r = section_header(ws, r, "Carousel Data Flow Notes", NC, C_HEADER_BLUE)
    notes = [
        "1. App sends a carousel request (opcode 0x07, motorId in SS) when the user swipes to a different carousel page.",
        "2. logFetchJob (background coroutine in each ViewModel) also cycles ALL motor IDs at 400ms intervals while logging is active,",
        "   ensuring all motors are captured in the CSV log from the first second, regardless of which page the user has visited.",
        "3. Draw Frame (DF): carousel data arrives EMBEDDED in opcode 0x06 run-state frames (Flutter protocol) in addition to separate 0x07 frames.",
        "4. Flyer Frame: carousel data ALSO arrives embedded in opcode 0x06 frames (same approach as DF).",
        "5. Carding and Ring: carousel data arrives as separate opcode 0x07 frames only.",
        "6. logFetchJob pauses when isDiagnosing=true to avoid interfering with diagnosis start/stop timing.",
    ]
    for i, note in enumerate(notes):
        write_row(ws, r, [note, "", "", "", "", "", ""], alts[i % 2])
        r += 1


# ─── SHEET 9 : PID ───────────────────────────────────────────────────────────
def sheet_pid(wb):
    ws = wb.create_sheet("PID")
    freeze(ws, "A3")
    set_col_widths(ws, [26, 16, 16, 20, 18, 18, 50])
    NC = 7
    alts = [C_ROW_BLUE_L, C_WHITE]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "PID Protocol — Common to All 4 Machines")
    c.fill = fill(C_HEADER_BLUE); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    r = section_header(ws, r, "PID Request  |  opcode 0x0D  |  substate 0x00  |  App → Board", NC, C_HEADER_BLUE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "", "", "Notes"], C_HEADER_BLUE)
    write_row(ws, r, ["Motor ID", "0x01", "02", "UInt8 (Compact)", "", "", "Which motor's PID to request. Same motor IDs as carousel/diagnosis"], C_ROW_BLUE_L, code_cols={1, 2, 3})
    r += 1

    r += 1
    r = section_header(ws, r, "PID Response  |  opcode 0x0E  |  substate 0x00  |  Board → App", NC, C_HEADER_BLUE)
    r = col_headers(ws, r, ["TLV Name", "Type (Hex)", "WireLen", "Data Type", "Unit", "Field", "Notes"], C_HEADER_BLUE)
    pid_resp = [
        ("Kp  (Proportional)",  "0x01", "08", "Float", "—", "Kp",  "Proportional gain"),
        ("Ki  (Integral)",      "0x02", "08", "Float", "—", "Ki",  "Integral gain"),
        ("FF  (Feed Forward)",  "0x03", "08", "Float", "—", "FF",  "Feed-forward term"),
        ("SO  (Speed Offset)",  "0x04", "08", "Float", "—", "SO",  "Speed offset term"),
    ]
    for i, row in enumerate(pid_resp):
        write_row(ws, r, row, alts[i % 2], code_cols={1, 2, 3})
        r += 1

    r += 1
    r = section_header(ws, r, "PID Update  |  opcode 0x0F  |  substate 0x00  |  App → Board  |  Same TLVs as response", NC, C_HEADER_BLUE)
    write_row(ws, r, ["TLV codes 0x01–0x04, Float, same structure as the 0x0E response. Sends updated Kp/Ki/FF/SO back to board.", "", "", "", "", "", ""], C_ROW_BLUE_L)
    r += 1

    r += 1
    r = section_header(ws, r, "PID Flow", NC, C_HEADER_GREEN)
    flow = [
        ("1. App sends opcode 0x0D with TLV 0x01 = motorId (UInt8)", ""),
        ("2. Board replies with opcode 0x0E, 4 float TLVs (Kp, Ki, FF, SO)", ""),
        ("3. User edits values on PID screen", ""),
        ("4. App sends opcode 0x0F with same 4 float TLVs containing updated values", ""),
    ]
    for i, (step, _) in enumerate(flow):
        write_row(ws, r, [step, "", "", "", "", "", ""], alts[i % 2])
        r += 1


# ─── SHEET 10 : Defaults & Limits ─────────────────────────────────────────────
def sheet_defaults(wb):
    ws = wb.create_sheet("Defaults & Limits")
    freeze(ws, "A3")
    set_col_widths(ws, [30, 18, 18, 18, 50])
    NC = 5
    alts = [C_ROW_BLUE_L, C_WHITE]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Default Settings & Valid Limits — All 4 Machines")
    c.fill = fill(C_HEADER_BLUE); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    def machine_block(title, bg_h, fields):
        nonlocal r
        r = section_header(ws, r, title, NC, bg_h)
        r = col_headers(ws, r, ["Field Name", "TLV (Hex)", "Default Value", "Valid Range", "Notes"], bg_h)
        for i, row in enumerate(fields):
            write_row(ws, r, row, alts[i % 2], code_cols={1})
            r += 1
        r += 1

    machine_block("Draw Frame (DF) — Settings Defaults & Limits", C_HEADER_GREEN, [
        ("Delivery Speed",        "0x70", "80",   "50 – 300",     "m/min (integer)"),
        ("Draft",                 "0x71", "8.0",  "3.0 – 12.0",   "Dimensionless float"),
        ("Length Limit",          "0x72", "400",  "5 – 1250",     "Metres"),
        ("Ramp Up Time",          "0x73", "6",    "3 – 15",       "Seconds"),
        ("Ramp Down Time",        "0x74", "6",    "3 – 15",       "Seconds"),
        ("Creel Tension Factor",  "0x75", "1.0",  "0.25 – 5.0",   "Float"),
    ])

    machine_block("Draw Frame — AL Settings Limits (opcode 0x11 TLVs)", C_HEADER_ORANGE, [
        ("KP Gain",            "0x20", "—",    "0 < KP ≤ 1.0",   "Float; must be positive and ≤ 1"),
        ("Sliver N+1 Thresh",  "0x21", "—",    "< Sliver N",     "Integer ADC counts; must be strictly less than N"),
        ("Sliver N Thresh",    "0x22", "—",    "< Sliver N-1",   "Integer ADC counts; must be between N+1 and N-1"),
        ("Sliver N-1 Thresh",  "0x23", "—",    "Largest value",  "Integer ADC counts; constraint: N+1 < N < N-1"),
        ("Target g/m",         "0x24", "—",    "4 < target ≤ 6", "Float grams per metre"),
    ])

    machine_block("Carding (Blow Card) — Settings Defaults & Limits", C_HEADER_GREEN, [
        ("Delivery Speed",        "0x80", "8.0",  "2.5 – 50.0",   "m/min (Float)"),
        ("Card Feed Ratio",       "0x81", "5.0",  "0.4 – 20.0",   "Float; used in derived RPM calc"),
        ("Length Limit",          "0x82", "1000", "100 – 1000",   "Metres"),
        ("Cylinder Speed",        "0x83", "750",  "300 – 875",    "RPM"),
        ("Beater Speed",          "0x84", "600",  "300 – 650",    "RPM"),
        ("Picker Cylinder Speed", "0x85", "600",  "300 – 700",    "RPM"),
        ("Beater Feed",           "0x86", "10",   "1 – 11",       "Int; ×120 = Beater Feed Motor RPM"),
        ("AF Feed",               "0x87", "7",    "1 – 8",        "Int; Auto-feed speed"),
    ])

    machine_block("Flyer Frame — Settings Defaults & Limits", "7F6000", [
        ("Spindle Speed",       "0x50", "650",  "500 – 1100",   "RPM"),
        ("Draft",               "0x51", "8.8",  "5.0 – 15.0",   "Float"),
        ("Twist Per Inch",      "0x52", "1.4",  "1.0 – 1.6",    "TPI Float"),
        ("RTF",                 "0x53", "1.0",  "0.5 – 1.5",    "Running Tension Factor"),
        ("Layers",              "0x54", "50",   "1 – 100",       "Count"),
        ("Max Height",          "0x55", "270",  "250 – 300",    "mm"),
        ("Roving Width",        "0x56", "1.2",  "1.0 – 8.0",    "mm Float"),
        ("Delta Bobbin Dia",    "0x57", "1.1",  "0.3 – 2.5",    "mm per layer"),
        ("Bare Bobbin Dia",     "0x58", "48",   "46 – 52",      "mm"),
        ("Ramp Up Time",        "0x59", "12",   "5 – 20",       "Seconds"),
        ("Ramp Down Time",      "0x60", "12",   "5 – 20",       "Seconds"),
        ("Change Layer Time",   "0x61", "800",  "200 – 2500",   "ms"),
        ("Cone Angle Factor",   "0x62", "1.0",  "0.1 – 3.0",    "Float"),
    ])

    machine_block("Ring Doubler — Settings Defaults & Limits", C_HEADER_PURPLE, [
        ("Input Yarn (Ne)",       "0x90", "30",    "10 – 80",    "Yarn count Ne"),
        ("Output Yarn Dia",       "0x91", "auto",  "0.05 – 1.5", "mm; blank → app auto-calculates"),
        ("Spindle Speed",         "0x92", "8000",  "6000 – 12000","RPM"),
        ("Twist Per Inch",        "0x93", "20",    "12 – 30",    "TPI integer"),
        ("Package Height",        "0x94", "150",   "50 – 250",   "mm"),
        ("Dia Build Factor",      "0x95", "0.5",   "0.05 – 2.5", "Float"),
        ("Winding Closeness",     "0x96", "100",   "50 – 200",   "Percent"),
        ("Winding Offset Coils",  "0x97", "2",     "1 – 5",      "Count"),
    ])


# ─── SHEET 11 : Quick Reference ──────────────────────────────────────────────
def sheet_quick_ref(wb):
    ws = wb.create_sheet("Quick Reference")
    freeze(ws, "A3")
    set_col_widths(ws, [22, 30, 20, 14, 30])
    NC = 5
    alts = [C_ROW_BLUE_L, C_WHITE]

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    c = ws.cell(r, 1, "Quick Reference — TLV Type Code Master Index")
    c.fill = fill(C_HEADER_BLUE); c.font = Font(bold=True, color="FFFFFF", size=14)
    c.alignment = center(); r += 1

    r = col_headers(ws, r, ["TLV Type (Hex)", "Name", "Data Type", "Machine(s)", "Context / Opcode"], C_HEADER_BLUE)

    all_tlvs = [
        # Common run-state / carousel TLVs (0x01–0x10)
        ("0x01", "Left Lift / Pause Reason / Error Info",    "Context-dependent",  "Flyer, Ring, DF, Carding", "opcode 0x06 (substate-dependent)"),
        ("0x02", "Right Lift / Pause Layer / Error Source",  "Context-dependent",  "Flyer, Ring, DF, Carding", "opcode 0x06 (substate-dependent)"),
        ("0x03", "Layer Count / Error Code",                 "UInt16",             "Flyer, Ring",              "opcode 0x06"),
        ("0x04", "Motor Temp / Error Layer",                 "UInt16",             "All",                      "opcode 0x06 / 0x07"),
        ("0x05", "MOSFET Temp",                              "UInt16",             "All",                      "opcode 0x06 / 0x07"),
        ("0x06", "Motor Current",                            "Float",              "All",                      "opcode 0x06 / 0x07"),
        ("0x07", "Motor RPM",                                "UInt16",             "All",                      "opcode 0x06 / 0x07"),
        ("0x08", "Output Metres",                            "Float",              "All",                      "opcode 0x06 / 0x07"),
        ("0x09", "Motor ID (what info)",                     "UInt8 (Compact)",    "All",                      "opcode 0x06 / 0x07 — identifies motor in carousel"),
        ("0x0A", "Total Power",                              "Float",              "All",                      "opcode 0x06 / 0x07"),
        ("0x0B", "Carding Duct Sensor",                      "UInt8 (Compact)",    "Carding",                  "opcode 0x06 running — 1=Open, 0=Blocked"),
        ("0x0C", "Coiler Sensor",                            "UInt8 (Compact)",    "Carding",                  "opcode 0x06 running — 1=Open, 0=Blocked"),
        ("0x0D", "Delivery Speed",                           "Float",              "DF, Carding",              "opcode 0x06 / 0x07 (m/min)"),
        ("0x0E", "Current Length",                           "Float",              "DF",                       "opcode 0x06 / 0x07 (metres)"),
        ("0x0F", "AL Sensor State / Front Cut Count",        "UInt16",             "DF / Flyer",               "DF 0x06 running: 1=Active. Flyer 0x06 running: cut count"),
        ("0x10", "Weight / Back Cut Count",                  "Float / UInt16",     "Ring / Flyer",             "Ring 0x06: weight(Float). Flyer 0x06: back cut count(UInt16)"),
        # DF AL Settings (0x20–0x24)
        ("0x20", "AL KP Gain",                              "Float",              "DF",                       "opcode 0x11 AL Save / 0x12 AL Response"),
        ("0x21", "AL Sliver N+1 Threshold",                 "UInt16",             "DF",                       "opcode 0x11 / 0x12"),
        ("0x22", "AL Sliver N Threshold",                   "UInt16",             "DF",                       "opcode 0x11 / 0x12"),
        ("0x23", "AL Sliver N-1 Threshold",                 "UInt16",             "DF",                       "opcode 0x11 / 0x12"),
        ("0x24", "AL Target g/m",                           "Float",              "DF",                       "opcode 0x11 / 0x12"),
        # DF Calibration (0x30–0x31)
        ("0x30", "Calibration Avg ADC",                     "UInt16",             "DF",                       "opcode 0x15 calibration result"),
        ("0x31", "Calibration Sample Count",                "UInt16",             "DF",                       "opcode 0x15 calibration result"),
        # Flyer Settings (0x50–0x62)
        ("0x50", "Spindle Speed",                           "UInt16",             "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x51", "Draft",                                   "Float",              "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x52", "Twist Per Inch",                          "Float",              "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x53", "RTF",                                     "Float",              "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x54", "Layers",                                  "UInt16",             "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x55", "Max Height",                              "UInt16",             "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x56", "Roving Width",                            "Float",              "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x57", "Delta Bobbin Dia",                        "Float",              "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x58", "Bare Bobbin Dia",                         "UInt16",             "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x59", "Ramp Up Time",                            "UInt16",             "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x60", "Ramp Down Time",                          "UInt16",             "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x61", "Change Layer Time",                       "UInt16",             "Flyer",                    "opcode 0x01 / 0x02"),
        ("0x62", "Cone Angle Factor",                       "Float",              "Flyer",                    "opcode 0x01 / 0x02"),
        # DF Settings (0x70–0x75)
        ("0x70", "Delivery Speed",                          "UInt16",             "DF",                       "opcode 0x01 / 0x02 settings"),
        ("0x71", "Draft",                                   "Float",              "DF",                       "opcode 0x01 / 0x02 settings"),
        ("0x72", "Length Limit",                            "UInt16",             "DF",                       "opcode 0x01 / 0x02 settings"),
        ("0x73", "Ramp Up Time",                            "UInt16",             "DF",                       "opcode 0x01 / 0x02 settings"),
        ("0x74", "Ramp Down Time",                          "UInt16",             "DF",                       "opcode 0x01 / 0x02 settings"),
        ("0x75", "Creel Tension Factor",                    "Float",              "DF",                       "opcode 0x01 / 0x02 settings"),
        # Carding Settings (0x80–0x87)
        ("0x80", "Delivery Speed",                          "Float",              "Carding",                  "opcode 0x01 / 0x02 settings"),
        ("0x81", "Card Feed Ratio",                         "Float",              "Carding",                  "opcode 0x01 / 0x02 settings"),
        ("0x82", "Length Limit",                            "UInt16",             "Carding",                  "opcode 0x01 / 0x02 settings"),
        ("0x83", "Cylinder Speed",                          "UInt16",             "Carding",                  "opcode 0x01 / 0x02 settings"),
        ("0x84", "Beater Speed",                            "UInt16",             "Carding",                  "opcode 0x01 / 0x02 settings"),
        ("0x85", "Picker Cylinder Speed",                   "UInt16",             "Carding",                  "opcode 0x01 / 0x02 settings"),
        ("0x86", "Beater Feed",                             "UInt16",             "Carding",                  "opcode 0x01 / 0x02 settings"),
        ("0x87", "AF Feed",                                 "UInt16",             "Carding",                  "opcode 0x01 / 0x02 settings"),
        # Ring Settings (0x90–0x97)
        ("0x90", "Input Yarn (Ne)",                         "UInt16",             "Ring",                     "opcode 0x01 / 0x02 settings"),
        ("0x91", "Output Yarn Diameter",                    "Float",              "Ring",                     "opcode 0x01 / 0x02 settings"),
        ("0x92", "Spindle Speed",                           "UInt16",             "Ring",                     "opcode 0x01 / 0x02 settings"),
        ("0x93", "Twist Per Inch",                          "UInt16",             "Ring",                     "opcode 0x01 / 0x02 settings"),
        ("0x94", "Package Height",                          "UInt16",             "Ring",                     "opcode 0x01 / 0x02 settings"),
        ("0x95", "Dia Build Factor",                        "Float",              "Ring",                     "opcode 0x01 / 0x02 settings"),
        ("0x96", "Winding Closeness",                       "UInt16",             "Ring",                     "opcode 0x01 / 0x02 settings"),
        ("0x97", "Winding Offset Coils",                    "UInt16",             "Ring",                     "opcode 0x01 / 0x02 settings"),
        # Diagnostics TLVs (0x40–0x45)
        ("0x40", "Motor ID (diagnosis)",                    "UInt8 (Compact)",    "All",                      "opcode 0x04 start command"),
        ("0x41", "Control Type",                            "UInt8 (Compact)",    "All",                      "opcode 0x04: 0x01=OpenLoop, 0x02=ClosedLoop"),
        ("0x42", "Speed Percent",                           "UInt16",             "All",                      "opcode 0x04: 0–100%"),
        ("0x43", "Duration (seconds)",                      "UInt16",             "All",                      "opcode 0x04: 1–30s"),
        ("0x44", "Direction",                               "UInt8 (Compact)",    "All",                      "opcode 0x04: 0x01=Fwd/Up, 0x02=Rev/Down"),
        ("0x45", "Bed Distance mm",                         "UInt16",             "Flyer, Ring",              "opcode 0x04: lift motors only — replaces Speed% and Duration"),
        # PID TLVs (0x01–0x04 in PID opcodes)
        ("0x01 (PID context)", "Motor ID (PID req) / Kp (PID resp/update)", "UInt8 / Float", "All", "opcode 0x0D: motor ID. opcode 0x0E/0x0F: Kp"),
        ("0x02 (PID context)", "Ki",                        "Float",              "All",                      "opcode 0x0E / 0x0F"),
        ("0x03 (PID context)", "FF (Feed Forward)",          "Float",              "All",                      "opcode 0x0E / 0x0F"),
        ("0x04 (PID context)", "SO (Speed Offset)",          "Float",              "All",                      "opcode 0x0E / 0x0F"),
    ]
    for i, row in enumerate(all_tlvs):
        write_row(ws, r, row, alts[i % 2], code_cols={0})
        r += 1


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    sheet_frame_format(wb)
    sheet_opcodes(wb)
    sheet_draw_frame(wb)
    sheet_carding(wb)
    sheet_flyer(wb)
    sheet_ring(wb)
    sheet_diagnostics(wb)
    sheet_carousel(wb)
    sheet_pid(wb)
    sheet_defaults(wb)
    sheet_quick_ref(wb)

    out = r"C:\Users\Jeeva\Desktop\App_Discussion\Gen4_Spinning_HMI_v8.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

main()
